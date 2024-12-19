import openpyxl
import os
import tools

 
directory_path = tools.get_path()
finalpath = tools.get_finalpath()

 
for filename in os.listdir(directory_path):
    if filename.endswith(".xlsx"):
        workbook_path = os.path.join(directory_path, filename)
        
        
        workbook = openpyxl.load_workbook(workbook_path)
 
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            if sheet_name == 'Store List':
                
                cell = f"DZ{2}"
                value = "."
                sheet[cell] = value
                sheet[cell].number_format = '@'
                
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = '@'
           
            
            if sheet_name == 'Store Summary':
                    cell = f"DU{1}"
                    value = "."
                    sheet[cell] = value
                    sheet[cell].number_format = '@'
                    sheet[cell] = value
 
        workbook.save(workbook_path)
        
        processed_directory = finalpath
        if not os.path.exists(processed_directory):
            os.makedirs(processed_directory)

        new_workbook_path = os.path.join(processed_directory, filename)
        os.replace(workbook_path, new_workbook_path)
